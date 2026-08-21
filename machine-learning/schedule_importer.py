"""
Turn the departmental schedule workbook into faculty and faculty_schedule rows.

    python machine-learning/schedule_importer.py --dry-run
    python machine-learning/schedule_importer.py --out database/sample-data/002_ccsict_schedule.sql

Reads `CCSICT-SCHEDULE-of-1st-SEm-2026-2027.xlsx` and emits SQL. It does NOT
write to the database: this is real, identifiable data about 37 people who have
not consented to anything, and it should be read by a human before it lands in
a shared Postgres instance.

WHAT THE WORKBOOK ACTUALLY IS. Eleven sheets in two shapes:

  `Faculty`   Echague, SY 2026-2027. 37 blocks, one per lecturer, each a
              Mon-Fri timetable grid. THE ONLY SOURCE OF ROWS. A lecturer's
              block holds every class they teach, and the ones taught in
              Santiago are marked STGO in the section, which is where the
              `campus` column comes from.
  `STGO`      Santiago City, SY **2025-2026** -- a PREVIOUS YEAR left in the
              workbook. Not imported and not compared against: reconciling two
              different school years reports every ordinary year-on-year change
              as a defect in the current data. See school_year_of().

`1ST`..`4TH` and the `ROOM`/`LAB` sheets are section- and room-centric views of
the same teaching, and are not read.

THREE THINGS THE FILE WILL DO TO A NAIVE PARSER:

1. Merged cells carry the duration. A 7:00-10:00 class is one merge spanning
   six 30-minute rows, with text in the first only. Reading cell-by-cell yields
   a 30-minute class and five empty rows. `read_only=True` silently discards
   merge geometry, so the workbook is opened normally.

2. Times have no meridiem. `1:00 - 1:30` is the afternoon. `strptime` puts it
   at one in the morning. Resolved by grid position instead: the teaching day
   runs 7:00 through 18:30, so 7:00-11:59 is AM and everything else is PM.

3. `DILAMANTA` is `DIMALANTA, OLIVER`, and one cell's room label sits where a
   surname belongs. Both are corrected loudly -- see ALIASES and the unmatched
   report. Silently dropping a lecturer's classes would make them look free.
"""

from __future__ import annotations

import argparse
import pathlib
import re
import sys
from collections import defaultdict
from datetime import time

import openpyxl

ROOT = pathlib.Path(__file__).resolve().parent.parent
WORKBOOK = ROOT / "CCSICT-SCHEDULE-of-1st-SEm-2026-2027.xlsx"

SEMESTER = "2026-2027-1"
DEPARTMENT = ("College of Computing, Information and Communication Technology",
              "CCSICT", "CCSICT")

# 0 = Sunday, matching faculty_schedule.day_of_week.
DAY_COLUMNS = {1: "MONDAY", 2: "TUESDAY", 3: "WEDNESDAY", 4: "THURSDAY", 5: "FRIDAY"}

# Surnames written differently in the STGO sheet than in the roster. Kept as an
# explicit table, and every application is printed, because a silent fuzzy match
# on a person's name is how one lecturer's classes end up on another's record.
ALIASES = {"DILAMANTA": "DIMALANTA"}

# A cell's middle line should be a section or a surname. These are rooms, and
# appear there only when the cell is missing a line.
ROOM_WORDS = re.compile(r"^(LAB|ROOM|RM|R)\s*\d*$|^(IT LAB|CENTRUM|RX)\b", re.I)


# ---------------------------------------------------------------- time parsing
def _parse_clock(token: str) -> time | None:
    """`7:00` -> 07:00, `1:30` -> 13:30. See docstring note 2."""
    m = re.match(r"^\s*(\d{1,2}):(\d{2})\s*$", token)
    if not m:
        return None
    h, mi = int(m.group(1)), int(m.group(2))
    if h < 7:                 # 1:00 .. 6:30 are afternoon on a 7am-6:30pm grid
        h += 12
    return time(h, mi)


def parse_time_range(text: str) -> tuple[time, time] | None:
    """`'7:00 - 7:30'` -> (07:00, 07:30)."""
    if not text or "-" not in str(text):
        return None
    a, _, b = str(text).partition("-")
    start, end = _parse_clock(a), _parse_clock(b)
    if not start or not end:
        return None
    return start, end


# ------------------------------------------------------------------ cell shape
def split_cell(raw: str) -> tuple[str, str, str]:
    """
    `'MA 215\\nBSMA 2-2  R112'`     -> ('MA 215', 'BSMA 2-2', 'R112')
    `'IT 111\\nBSIT 1-1 NS STGO\\nLAB 2'` -> ('IT 111', 'BSIT 1-1 NS STGO', 'LAB 2')
    """
    lines = [l.strip() for l in str(raw).split("\n") if l.strip()]
    if not lines:
        return "", "", ""
    course = lines[0]
    middle = lines[1] if len(lines) > 1 else ""
    room = lines[2] if len(lines) > 2 else ""

    # A room tacked onto the end of the section line, two spaces in.
    if not room and middle:
        m = re.search(r"\s{2,}(\S.*)$", middle)
        if m:
            room = m.group(1).strip()
            middle = middle[: m.start()].strip()
    return course, middle, room


def campus_of(*parts: str) -> str:
    """Santiago blocks are marked STG or STGO somewhere in the cell."""
    blob = " ".join(p.upper() for p in parts if p)
    return "santiago" if re.search(r"\bSTGO?\b", blob) else "echague"


# ------------------------------------------------------------------- traversal
def _merge_lookup(ws):
    """(row, col) 1-based -> (row_span, is_origin) for merged regions."""
    spans = {}
    for rng in ws.merged_cells.ranges:
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                spans[(r, c)] = (rng.min_row, rng.max_row, r == rng.min_row and c == rng.min_col)
    return spans


def parse_faculty_sheet(ws) -> tuple[list[str], list[dict]]:
    """The authoritative pass: every lecturer, every block, both campuses."""
    rows = list(ws.iter_rows(values_only=True))
    spans = _merge_lookup(ws)
    names, blocks = [], []

    starts = [i for i, r in enumerate(rows)
              if r and r[0] and str(r[0]).strip().upper().startswith("NAME")]

    for i in starts:
        name = next((str(c).strip() for c in rows[i][1:] if c and str(c).strip()), "")
        if not name:
            continue
        names.append(name)

        header = next((j for j in range(i, min(i + 12, len(rows)))
                       if rows[j][0] and str(rows[j][0]).strip().upper() == "TIME"), None)
        if header is None:
            continue
        stop = next((j for j in range(header + 1, len(rows))
                     if rows[j][0] and str(rows[j][0]).strip().upper().startswith(("CODE", "NAME"))),
                    len(rows))

        for j in range(header + 1, stop):
            span = parse_time_range(rows[j][0])
            if not span:
                continue
            for dow, _label in DAY_COLUMNS.items():
                value = rows[j][dow]
                if not value or not str(value).strip():
                    continue

                # Only the origin of a merge carries the text; its END row is
                # what gives the class its real duration.
                merged = spans.get((j + 1, dow + 1))
                if merged:
                    min_row, max_row, is_origin = merged
                    if not is_origin:
                        continue
                    last = parse_time_range(rows[max_row - 1][0])
                    end = last[1] if last else span[1]
                else:
                    end = span[1]

                course, middle, room = split_cell(value)
                blocks.append({
                    "faculty": name, "dow": dow,
                    "start": span[0], "end": end,
                    "course": course, "section": middle, "room": room,
                    "campus": campus_of(course, middle, room),
                })
    return names, blocks


def school_year_of(ws) -> str | None:
    """
    The `SCHOOL YEAR:` a sheet declares, from its first few rows.

    Worth reading because the tabs in this workbook are NOT all the same year.
    `Faculty` is Echague SY 2026-2027; `STGO`, `ROOM STGO` and `LAB STGO` are
    Santiago SY 2025-2026, left over from a previous file. Reconciling one
    against the other compares two different semesters and reports every
    ordinary year-on-year change as a data defect.
    """
    for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
        for i, c in enumerate(row):
            if c and "SCHOOL YEAR" in str(c).upper():
                tail = " ".join(str(x) for x in row[i:] if x)
                m = re.search(r"(\d{4})\s*-\s*(\d{4})", tail)
                if m:
                    return f"{m.group(1)}-{m.group(2)}"
        # Some sheets put the year in a free-text banner instead of a label.
        for c in row:
            if c and "SEMESTER" in str(c).upper():
                m = re.search(r"(\d{4})\s*-\s*(\d{4})", str(c))
                if m:
                    return f"{m.group(1)}-{m.group(2)}"
    return None


def stated_contact_hours(ws) -> dict[str, float]:
    """
    Each block's own `CONTACT HOURS :` figure, where it is filled in.

    A free correctness check the document gives away: if the parsed grid does
    not add up to the number the department wrote at the top of the block, one
    of the two is wrong, and it is worth knowing which before the rows reach
    a database. Most blocks leave it blank, so this covers a sample rather than
    the roster.
    """
    rows = list(ws.iter_rows(values_only=True))
    out = {}
    for i, r in enumerate(rows):
        if not (r and r[0] and str(r[0]).strip().upper().startswith("NAME")):
            continue
        name = next((str(c).strip() for c in r[1:] if c and str(c).strip()), "")
        if not name:
            continue
        for j in range(i, min(i + 6, len(rows))):
            for k, c in enumerate(rows[j]):
                if c and "CONTACT" in str(c).upper():
                    v = next((x for x in rows[j][k + 1:] if isinstance(x, (int, float))), None)
                    if v is not None:
                        out[name] = float(v)
    return out


def parse_stgo_sheet(ws, roster: list[str]) -> tuple[list[dict], list[str]]:
    """Cross-check pass. Returns (blocks, unmatched-cell reports)."""
    rows = list(ws.iter_rows(values_only=True))
    spans = _merge_lookup(ws)
    by_surname = {n.split(",")[0].strip().upper(): n for n in roster}
    found, unmatched = [], []

    header_rows = [i for i, r in enumerate(rows)
                   if r and r[0] and str(r[0]).strip().upper() == "TIME"]

    for h in header_rows:
        stop = next((j for j in range(h + 1, len(rows))
                     if rows[j][0] and str(rows[j][0]).strip().upper() == "TIME"), len(rows))
        for j in range(h + 1, stop):
            span = parse_time_range(rows[j][0])
            if not span:
                continue
            for dow in DAY_COLUMNS:
                value = rows[j][dow] if dow < len(rows[j]) else None
                if not value or not str(value).strip():
                    continue
                merged = spans.get((j + 1, dow + 1))
                if merged:
                    min_row, max_row, is_origin = merged
                    if not is_origin:
                        continue
                    last = parse_time_range(rows[max_row - 1][0])
                    end = last[1] if last else span[1]
                else:
                    end = span[1]

                course, middle, room = split_cell(value)
                key = ALIASES.get(middle.upper(), middle.upper())

                if key not in by_surname:
                    # A room label where a surname belongs: the cell is missing
                    # its lecturer line. Report it; do not guess.
                    why = "room label in the name position" if ROOM_WORDS.match(middle) \
                          else "no roster match"
                    unmatched.append(f"{ws.title} row {j+1} {DAY_COLUMNS[dow]}: "
                                     f"{course!r} / {middle!r} -- {why}")
                    continue

                if middle.upper() in ALIASES:
                    print(f"  alias applied: {middle!r} -> {by_surname[key]!r} "
                          f"({ws.title} row {j+1})")

                found.append({
                    "faculty": by_surname[key], "dow": dow,
                    "start": span[0], "end": end,
                    "course": course, "section": "", "room": room,
                    "campus": "santiago",
                })
    return found, unmatched


# ------------------------------------------------------------------------- SQL
def q(s: str) -> str:
    return "'" + str(s).replace("'", "''") + "'"


def emit_sql(names: list[str], blocks: list[dict]) -> str:
    out = [
        "-- Generated by machine-learning/schedule_importer.py -- do not hand-edit.",
        f"-- Source: {WORKBOOK.name}",
        f"-- Semester: {SEMESTER}   Faculty: {len(names)}   Blocks: {len(blocks)}",
        "--",
        "-- data_origin='real': these are real institutional records.",
        "-- is_consented stays false (schema default). The assistant will refuse",
        "-- every availability question about these people until written consent",
        "-- exists. That is the designed behaviour, not a bug to work around.",
        "",
        "begin;",
        "set search_path = geobot, public;",
        "",
        "-- 1. Department",
        "insert into department (name, short_code, college, data_origin)",
        f"values ({q(DEPARTMENT[0])}, {q(DEPARTMENT[1])}, {q(DEPARTMENT[2])}, 'real')",
        "on conflict (name) do nothing;",
        "",
        "-- 2. Faculty. faculty has no unique constraint on full_name, so the",
        "--    guard is an explicit NOT EXISTS rather than ON CONFLICT.",
        "insert into faculty (full_name, department_id, data_origin)",
        "select v.full_name, d.id, 'real'",
        "  from (values",
    ]
    out.append(",\n".join(f"    ({q(n)})" for n in sorted(names)))
    out += [
        "  ) as v(full_name)",
        f"  cross join (select id from department where short_code = {q(DEPARTMENT[1])}) d",
        " where not exists (select 1 from faculty f where f.full_name = v.full_name);",
        "",
        "-- 3. Pseudonyms. The ML feature store keys on these and never on a name",
        "--    (audit F-19); the column default generates them.",
        "insert into faculty_pseudonym_map (faculty_id)",
        "select f.id from faculty f",
        " where not exists (select 1 from faculty_pseudonym_map m where m.faculty_id = f.id);",
        "",
        "-- 4. Surname aliases, so the router's gazetteer resolves 'Prof. Alado'.",
        "insert into faculty_alias (faculty_id, alias, alias_kind, data_origin)",
        "select f.id, split_part(f.full_name, ',', 1), 'surname', 'real'",
        "  from faculty f",
        " where not exists (",
        "   select 1 from faculty_alias a",
        "    where a.faculty_id = f.id and a.alias = split_part(f.full_name, ',', 1));",
        "",
        "-- 5. Schedule. Replaced wholesale for this semester so a re-run is a",
        "--    reload rather than a duplication.",
        f"delete from faculty_schedule where semester = {q(SEMESTER)};",
        "",
        "insert into faculty_schedule",
        "  (faculty_id, day_of_week, start_time, end_time, block_kind,",
        "   semester, course_code, room_label, campus, data_origin)",
        "select f.id, v.dow, v.start_time, v.end_time, 'class',",
        f"       {q(SEMESTER)}, v.course, nullif(v.room, ''), v.campus, 'real'",
        "  from (values",
    ]
    rows = [
        f"    ({q(b['faculty'])}, {b['dow']}, "
        f"{q(b['start'].strftime('%H:%M'))}::time, {q(b['end'].strftime('%H:%M'))}::time, "
        f"{q(b['course'])}, {q(b['room'])}, {q(b['campus'])})"
        for b in blocks
    ]
    out.append(",\n".join(rows))
    out += [
        "  ) as v(full_name, dow, start_time, end_time, course, room, campus)",
        "  join faculty f on f.full_name = v.full_name;",
        "",
        "commit;",
        "",
    ]
    return "\n".join(out)


# ------------------------------------------------------------------------ main
def main():
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[1])
    ap.add_argument("--workbook", type=pathlib.Path, default=WORKBOOK)
    ap.add_argument("--out", type=pathlib.Path)
    ap.add_argument("--dry-run", action="store_true",
                    help="parse and report; write nothing")
    args = ap.parse_args()

    if not args.workbook.exists():
        raise SystemExit(f"workbook not found: {args.workbook}")

    wb = openpyxl.load_workbook(args.workbook, data_only=True)   # merges needed
    print(f"reading {args.workbook.name}")

    fac_year = school_year_of(wb["Faculty"])
    stgo_year = school_year_of(wb["STGO"])

    names, blocks = parse_faculty_sheet(wb["Faculty"])
    print(f"\n'Faculty' sheet: Echague, SY {fac_year or '?'} -- "
          f"{len(names)} lecturers, {len(blocks)} blocks")

    # ------------------------------------------------------------------
    # Reconcile only WITHIN a school year.
    #
    # The Santiago tabs in this workbook are a previous year's file left in
    # place. Cross-checking the current Echague schedule against them reports
    # every ordinary year-on-year change -- a class moved by thirty minutes, a
    # subject reassigned -- as a discrepancy in the current data, which is
    # noise dressed up as a finding. The campus tagging does not depend on
    # those tabs: it comes from the STGO markers inside this year's own sheet.
    # ------------------------------------------------------------------
    if stgo_year and fac_year and stgo_year != fac_year:
        print(f"'STGO' sheet:    Santiago, SY {stgo_year} -- SKIPPED")
        print(f"                 a different school year from 'Faculty' "
              f"(SY {fac_year}); not comparable, and not imported.")
        stgo, unmatched = [], []
    else:
        stgo, unmatched = parse_stgo_sheet(wb["STGO"], names)

    # ------------------------------------------------------------------
    # RECONCILIATION ONLY. The STGO sheet contributes no rows.
    #
    # It is a second VIEW of teaching the Faculty sheet already records, and
    # the two views disagree about times. DIMALANTA's Thursday is 07:00-09:00,
    # 09:00-11:00, 11:00-14:00 in STGO and 07:30-09:30, 09:30-11:30, 12:00-15:00
    # in Faculty -- the same three classes, half an hour apart.
    #
    # Importing the difference would not add classes, it would double-book
    # eleven lecturers. A person cannot be in two rooms at once, and a schedule
    # that says otherwise produces attendance labels that are wrong in a way no
    # later stage can detect. So the Faculty sheet wins, and every disagreement
    # is printed for someone to take back to the department.
    # ------------------------------------------------------------------
    have = {(b["faculty"], b["dow"], b["start"], b["end"]) for b in blocks}
    disputed = [b for b in stgo if (b["faculty"], b["dow"], b["start"], b["end"]) not in have]

    by_campus = defaultdict(int)
    for b in blocks:
        by_campus[b["campus"]] += 1

    if stgo:
        print(f"'STGO' sheet:    {len(stgo)} blocks cross-checked against 'Faculty'")
        print(f"                 {len(stgo) - len(disputed)} agree exactly, "
              f"{len(disputed)} disagree (NOT imported)")

    if disputed:
        print("\nSTGO entries the 'Faculty' sheet records differently:")
        for b in sorted(disputed, key=lambda x: (x["faculty"], x["dow"], x["start"])):
            clash = [o for o in blocks
                     if o["faculty"] == b["faculty"] and o["dow"] == b["dow"]
                     and not (o["end"] <= b["start"] or o["start"] >= b["end"])]
            note = ("Faculty sheet has "
                    + ", ".join(f"{o['start']:%H:%M}-{o['end']:%H:%M} {o['course']}"
                                for o in clash)) if clash else "no counterpart -- possibly missing"
            print(f"  {b['faculty'][:26]:<26} {DAY_COLUMNS[b['dow']]:<9} "
                  f"{b['start']:%H:%M}-{b['end']:%H:%M} {b['course']:<13} | {note}")
        print("\n  Worth confirming with the department which sheet is correct.")

    if unmatched:
        print(f"\n{len(unmatched)} STGO cell(s) skipped -- review by hand:")
        for u in unmatched:
            print(f"  {u}")

    print("\ntotals")
    print(f"  faculty          {len(names)}")
    print(f"  schedule blocks  {len(blocks)}")
    for c, n in sorted(by_campus.items()):
        print(f"    {c:<14} {n}")
    commuters = {b["faculty"] for b in blocks if b["campus"] == "santiago"}
    print(f"  teach on both    {len(commuters)}")

    # Cross-check against the department's own arithmetic.
    hours = defaultdict(float)
    for b in blocks:
        if b["campus"] == "echague" or True:      # stated totals cover both campuses
            hours[b["faculty"]] += (
                (b["end"].hour * 60 + b["end"].minute)
                - (b["start"].hour * 60 + b["start"].minute)
            ) / 60
    stated = stated_contact_hours(wb["Faculty"])
    agree = [n for n, v in stated.items() if abs(hours[n] - v) < 0.01]
    differ = [(n, hours[n], v) for n, v in stated.items() if abs(hours[n] - v) >= 0.01]
    print(f"\ncontact-hours cross-check ({len(stated)} of {len(names)} blocks state a total)")
    print(f"  parsed grid agrees   {len(agree)}")
    if differ:
        print(f"  disagrees            {len(differ)}")
        for n, got, want in differ:
            print(f"    {n:<34} grid={got:5.1f}  stated={want:5.1f}")
        print("  A disagreement is usually the source document, not the parser --"
              "\n  check the block's own subject legend before assuming otherwise.")

    bad = [b for b in blocks if b["end"] <= b["start"]]
    if bad:
        print(f"\n{len(bad)} block(s) end at or before they start -- refusing to emit SQL")
        for b in bad[:5]:
            print(f"  {b['faculty']} {DAY_COLUMNS[b['dow']]} {b['start']}-{b['end']} {b['course']}")
        sys.exit(1)

    if args.dry_run:
        print("\ndry run -- no SQL written")
        return

    sql = emit_sql(names, blocks)
    out = args.out or (ROOT / "database" / "sample-data" / "002_ccsict_schedule.sql")
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(sql, encoding="utf-8")
    print(f"\nwrote {out}  ({len(sql.splitlines())} lines)")
    print("Review it, then run it in the Supabase SQL editor.")


if __name__ == "__main__":
    main()

"""
Generate synthetic intraday attendance for the CCSICT roster.

    python machine-learning/generate_synthetic_attendance.py --dry-run
    python machine-learning/generate_synthetic_attendance.py

Writes two things that must agree with each other, from one run:

  database/sample-data/003_synthetic_attendance.sql
  machine-learning/training-data/synthetic-attendance-1st-sem-2026-2027.xlsx

WHY THIS EXISTS. The thesis advisor ruled that real Daily Time Records cannot
be requested -- a privacy concern -- and recommended synthetic data. Without
attendance of some kind there is no model to report: with schedule-derived
features AND schedule-derived labels the forest reproduces
schedule_lookup_status() by construction and cannot beat schedule_rule_baseline
(see feature_engineering.py).

WHAT IT CAN AND CANNOT SUPPORT. The generator injects per-lecturer behavioural
traits that are NOT derivable from a timetable -- how early someone arrives,
how often they are absent, whether they linger after the last class. Those
traits are the thing the classifier has to recover. So an accuracy measured on
this data answers:

    "does the pipeline recover patterns that were injected?"

and NEVER:

    "does the system predict real faculty availability?"

Every row is written with data_origin='synthetic', which makes
corpus_is_research_ready() return false, which makes train_availability_model.py
refuse to persist a reportable metric unless it is asked for a --simulation run.
That refusal is the design working, not an obstacle to route around.

NOBODY'S ATTENDANCE, LITERALLY. The cohort is 37 synthetic lecturers, SIM-01 to
SIM-37, carrying the real teaching SHAPES from the departmental workbook and
none of the real identities. See sim_name() for why that is not squeamishness:
load_roster() requires is_consented, no real lecturer has consented, and the
alternative to a synthetic cohort is writing a consent record that never
happened.

The real 37 stay in the database with their real schedules, is_consented=false
and no attendance at all -- production data, waiting on consent forms.
"""

from __future__ import annotations

import argparse
import hashlib
import pathlib
from datetime import date, datetime, time, timedelta

import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

import database_connector as db
import schedule_importer as si

ROOT = pathlib.Path(__file__).resolve().parent.parent

SEMESTER = "2026-2027-1"

# Reproducibility is a methodology requirement, not a convenience: the dataset
# has to be regenerable from this file alone for anyone checking the work.
SEED = 20260820
PSEUDONYM_SALT = "isu-geobot-ccsict-2026-2027-1"

# Travel between Echague and Santiago. A Santiago class does not merely occupy
# its own hours; it removes the surrounding ones from the Echague campus too.
TRAVEL_MARGIN = timedelta(minutes=90)

# THE CALENDAR IS READ, NOT DECLARED.
#
# This file used to carry its own holidays, its own exam periods and its own
# semester bounds, and it INSERTED them into institutional_event. That made the
# generator a second source of truth for the academic calendar, and it was the
# wrong one: the exam periods were placeholders invented before the official ISU
# calendar was consulted, and they were about a month adrift from the real ones.
#
# The authoritative rows now live in institutional_event, corrected against
# https://isu.edu.ph/school-calendar/ by 006_official_calendar.sql, and are read
# from there by the availability service, by dataset_loader, and by this file.
# One calendar, three readers.
def _load_calendar():
    """(window_start, window_end, {date: title} holidays, {date: title} exams)."""
    rows = db.fetch_all(
        "select event_date, event_type, title, disrupts_schedule "
        "from geobot.institutional_event order by event_date"
    )
    window = [r for r in rows
              if r["event_type"] == "other" and str(r["title"]).startswith("Academic window")]
    if len(window) != 2:
        raise SystemExit(
            "No academic window in institutional_event. Run "
            "database/sample-data/006_official_calendar.sql first -- the "
            "semester bounds are not this script's to invent."
        )
    holidays = {r["event_date"]: r["title"] for r in rows if r["event_type"] == "holiday"}
    exams = {r["event_date"]: r["title"] for r in rows if r["event_type"] == "exam_period"}
    return window[0]["event_date"], window[1]["event_date"], holidays, exams


SEM_START, SEM_END, HOLIDAYS, EXAM_DAYS = _load_calendar()


# ------------------------------------------------------------------ behaviour
class Traits:
    """
    One lecturer's habits, drawn once and then fixed for the semester.

    These are the injected ground truth. A classifier given only day, time and
    schedule cannot derive them, which is the entire reason the exercise
    produces something the rule baseline does not already do.
    """

    __slots__ = ("punctuality", "spread", "absence", "early_departure",
                 "consultation", "stay_late")

    def __init__(self, rng: np.random.Generator):
        # Minutes relative to the first class. Negative is early.
        self.punctuality = float(rng.normal(-12, 10))
        self.spread = float(rng.uniform(4, 14))
        self.absence = float(rng.beta(1.6, 26))          # ~2-10% of teaching days
        self.early_departure = float(rng.beta(2.2, 4.5))  # leaves at last bell
        self.consultation = float(rng.beta(3.2, 2.6))     # around on non-teaching hours
        self.stay_late = float(rng.uniform(10, 75))       # minutes after last class


def pseudonym_for(name: str) -> str:
    """Stable 18-hex surrogate, matching the shape of the column default."""
    return hashlib.sha256((PSEUDONYM_SALT + "|" + name).encode()).hexdigest()[:18]


def sim_name(index: int) -> str:
    """
    The simulation cohort's stand-in identity.

    WHY THE COHORT IS SYNTHETIC PEOPLE AND NOT THE REAL ROSTER.

    Two independent reasons, either sufficient on its own:

      Consent. load_roster() selects `where f.is_active and f.is_consented`,
      and every real lecturer is is_consented = false because none of them has
      signed anything. The only ways to train on the real roster are to obtain
      consent or to write a consent record that did not happen. The second is
      not an option -- a fabricated consent row is a false compliance claim
      sitting in a research database, and it would survive into the thesis.

      Attribution. The behaviour here is invented. Attaching an invented punch
      card to a named, identifiable colleague is the exact harm the pseudonym
      map exists to prevent, and it is not made acceptable by the data being
      fake -- if anything the reverse.

    So the cohort mirrors the real teaching SHAPES (loads, gaps, the Echague /
    Santiago split, which are properties of a timetable rather than of a
    person) and carries none of the identities. The real roster stays in the
    database, unconsented and without attendance, ready for the day consent
    forms exist.
    """
    return f"SIM-{index:02d}"


# -------------------------------------------------------------------- helpers
def weekdays(start: date, end: date):
    d = start
    while d <= end:
        if d.weekday() < 5:          # Mon..Fri
            yield d
        d += timedelta(days=1)


def exam_days() -> dict[date, str]:
    """Examination days, straight from institutional_event."""
    return {d: t for d, t in EXAM_DAYS.items() if d.weekday() < 5}


def at(day: date, t: time) -> datetime:
    return datetime.combine(day, t)


def merge_windows(spans: list[tuple[datetime, datetime]]) -> list[tuple[datetime, datetime]]:
    """Collapse overlapping/touching intervals."""
    if not spans:
        return []
    spans = sorted(spans)
    out = [spans[0]]
    for s, e in spans[1:]:
        if s <= out[-1][1]:
            out[-1] = (out[-1][0], max(out[-1][1], e))
        else:
            out.append((s, e))
    return out


# ------------------------------------------------------------------ generator
def generate(blocks: list[dict], names: list[str], rng: np.random.Generator):
    """Returns (punch rows, per-faculty traits, calendar rows)."""
    by_faculty_day: dict[tuple[str, int], list[dict]] = {}
    for b in blocks:
        by_faculty_day.setdefault((b["faculty"], b["dow"]), []).append(b)

    traits = {n: Traits(rng) for n in sorted(names)}
    exams = exam_days()
    punches: list[dict] = []

    for day in weekdays(SEM_START, SEM_END):
        if day in HOLIDAYS:
            continue
        # 0=Sunday in faculty_schedule; Python's Monday=0.
        dow = (day.weekday() + 1) % 7
        is_exam = day in exams

        for name in sorted(names):
            t = traits[name]
            todays = by_faculty_day.get((name, dow), [])

            echague = [b for b in todays if b["campus"] == "echague"]
            santiago = [b for b in todays if b["campus"] == "santiago"]

            # Hours the person cannot possibly be on this campus: the Santiago
            # class itself, plus the drive either side.
            away = merge_windows([
                (at(day, b["start"]) - TRAVEL_MARGIN, at(day, b["end"]) + TRAVEL_MARGIN)
                for b in santiago
            ])

            if not echague:
                # No local teaching at all. Someone may still come in for
                # consultation -- unless they are in Santiago, or it is an
                # exam week, when nobody drops by casually.
                if santiago or is_exam:
                    continue
                if rng.random() > t.consultation * 0.35:
                    continue
                start = at(day, time(9, 0)) + timedelta(minutes=int(rng.normal(0, 45)))
                end = start + timedelta(minutes=int(rng.uniform(90, 260)))
                punches.append({"name": name, "ts": start, "kind": "check_in"})
                punches.append({"name": name, "ts": end, "kind": "check_out"})
                continue

            if rng.random() < t.absence:
                continue

            # SEGMENTS, not one interval per day.
            #
            # A commuter's Echague day can be broken in half by a Santiago
            # class. Modelling the day as a single check-in/check-out either
            # deletes the afternoon or pretends they never left.
            #
            # And a block that starts inside the travel window is NOT dropped:
            # FERNANDEZ teaches Santiago until 13:00 and Echague at 14:00 on a
            # Wednesday, which 90 minutes of driving makes impossible to reach
            # on time. He arrives late. Dropping the block instead would erase a
            # lecturer from the dataset entirely, and lateness forced by a
            # timetable is exactly the kind of pattern a schedule lookup cannot
            # predict and a classifier can.
            free: list[tuple[datetime, datetime]] = []
            cursor_t = at(day, time(0, 0))
            for s, e in away:
                if s > cursor_t:
                    free.append((cursor_t, s))
                cursor_t = max(cursor_t, e)
            free.append((cursor_t, at(day, time(23, 59))))

            for lo, hi in free:
                here = [b for b in echague
                        if at(day, b["start"]) < hi and at(day, b["end"]) > lo]
                if not here:
                    continue

                first = min(at(day, b["start"]) for b in here)
                last = max(at(day, b["end"]) for b in here)

                arrive = first + timedelta(minutes=float(rng.normal(t.punctuality, t.spread)))
                if is_exam:
                    arrive = first - timedelta(minutes=float(rng.uniform(5, 25)))
                    depart = last + timedelta(minutes=float(rng.uniform(0, 20)))
                elif rng.random() < t.early_departure:
                    depart = last + timedelta(minutes=float(rng.uniform(0, 12)))
                else:
                    depart = last + timedelta(minutes=float(rng.normal(t.stay_late, 18)))

                # Clamp into the window they can actually be on campus.
                arrive = max(arrive, lo)
                depart = min(depart, hi)
                if depart <= arrive:
                    continue

                punches.append({"name": name, "ts": arrive, "kind": "check_in"})
                punches.append({"name": name, "ts": depart, "kind": "check_out"})

    punches.sort(key=lambda p: (p["name"], p["ts"]))

    calendar = [{"date": d, "type": "holiday", "title": h, "disrupts": True}
                for d, h in sorted(HOLIDAYS.items())]
    calendar += [{"date": d, "type": "exam_period", "title": t2, "disrupts": True}
                 for d, t2 in sorted(exams.items())]
    return punches, traits, calendar


# ------------------------------------------------------------------------ SQL
def emit_sql(punches, traits, calendar, names, blocks) -> str:
    q = si.q
    # Real name -> cohort stand-in. Sorted, so the mapping is stable
    # across runs and the spreadsheet agrees with the database.
    sim = {n: sim_name(i) for i, n in enumerate(sorted(names), start=1)}
    out = [
        "-- Generated by machine-learning/generate_synthetic_attendance.py",
        "-- DO NOT HAND-EDIT. Regenerate instead; the run is seeded and reproducible.",
        "--",
        "-- SYNTHETIC. Every row here is invented. It exists because the advisor",
        "-- ruled that real Daily Time Records cannot be requested (privacy), and",
        "-- because a forest trained on schedule-derived labels alone cannot beat",
        "-- the rule baseline it is supposed to be compared against.",
        "--",
        "-- data_origin='synthetic' is what makes corpus_is_research_ready() return",
        "-- false and stops these rows reaching a reported result by accident.",
        f"-- Seed {SEED}.  Semester {SEMESTER}.  {len(punches):,} punches.",
        "",
        "begin;",
        "set search_path = geobot, public;",
        "",
        "-- 1. The simulation cohort.",
        "--",
        "--    SYNTHETIC PEOPLE, not the real roster. load_roster() requires",
        "--    is_consented, no real lecturer has consented, and writing a consent",
        "--    row that never happened would put a false compliance claim in a",
        "--    research database. Inventing attendance for a named colleague is",
        "--    also the harm the pseudonym map exists to prevent.",
        "--",
        "--    These carry the real teaching SHAPES and none of the identities.",
        "--    is_consented is true because a simulated subject is in the study by",
        "--    construction; consent_date satisfies consent_requires_date.",
        "delete from attendance_record where data_origin = 'synthetic';",
        "delete from faculty_schedule where data_origin = 'synthetic';",
        "delete from faculty where data_origin = 'synthetic' and full_name like 'SIM-%';",
        "",
        "insert into faculty (full_name, department_id, is_consented, consent_date, data_origin)",
        "select v.full_name, d.id, true, date '2026-08-10', 'synthetic'",
        "  from (values",
    ]
    sims = [sim[n] for n in sorted(names)]
    out.append(",\n".join(f"    ({q(s)})" for s in sims))
    out += [
        "  ) as v(full_name)",
        "  cross join (select id from department where short_code = 'CCSICT') d;",
        "",
        "-- 2. Pseudonyms for the cohort. Deterministic, so the spreadsheet and",
        "--    the database agree without a lookup.",
        "insert into faculty_pseudonym_map (faculty_id, pseudonym_id)",
        "select f.id, v.pid",
        "  from (values",
    ]
    out.append(",\n".join(f"    ({q(s)}, {q(pseudonym_for(s))})" for s in sims))
    out += [
        "  ) as v(full_name, pid)",
        "  join faculty f on f.full_name = v.full_name",
        "on conflict (faculty_id) do nothing;",
        "",
        "-- 2. Academic calendar: DELIBERATELY NOT WRITTEN HERE.",
        "--",
        "--    institutional_event is authoritative, and 006_official_calendar.sql",
        "--    corrects it against https://isu.edu.ph/school-calendar/. This script",
        "--    now READS that calendar (see _load_calendar) instead of declaring",
        "--    one. It used to insert its own placeholder holidays and examination",
        "--    periods, which made it a competing source of truth -- and the wrong",
        "--    one, adrift by about a month on both examination windows.",
        "--",
        f"--    Window in force for this run: {SEM_START} to {SEM_END}",
        f"--    Read from the calendar: {len(HOLIDAYS)} holidays, {len(EXAM_DAYS)} examination days",
        "",
        "-- 4. The cohort's timetable: the real teaching shapes, on synthetic",
        "--    people. Marked synthetic so corpus_is_research_ready() counts it.",
        "insert into faculty_schedule",
        "  (faculty_id, day_of_week, start_time, end_time, block_kind,",
        "   semester, course_code, room_label, campus, data_origin)",
        "select f.id, v.dow, v.st, v.et, 'class',",
        f"       {q(SEMESTER)}, v.course, nullif(v.room, ''), v.campus, 'synthetic'",
        "  from (values",
    ]
    out.append(",\n".join(
        f"    ({q(sim[b['faculty']])}, {b['dow']}, "
        f"{q(b['start'].strftime('%H:%M'))}::time, {q(b['end'].strftime('%H:%M'))}::time, "
        f"{q(b['course'])}, {q(b['room'])}, {q(b['campus'])})"
        for b in blocks))
    out += [
        "  ) as v(full_name, dow, st, et, course, room, campus)",
        "  join faculty f on f.full_name = v.full_name;",
        "",
        "-- 5. Punches. Intraday pairs -- dataset_loader.py refuses anything else,",
        "--    because a daily sign-in sheet cannot produce intra-day labels",
        "--    without imputing them from the schedule, which is circular (F-18).",
        "insert into attendance_record",
        "  (pseudonym_id, event_time, event_type, source, granularity, data_origin)",
        "-- CAMPUS-LOCAL WALL CLOCK, converted explicitly.",
        "--",
        "-- event_time is timestamptz. A naive literal is read in the SESSION's",
        "-- timezone, which on this database is UTC, so a 07:00 check-in silently",
        "-- became 15:00 in Manila and every punch landed eight hours after the",
        "-- class it belongs to. The generator works in campus wall-clock time,",
        "-- so the conversion has to be stated rather than inherited.",
        "select m.pseudonym_id,",
        "       (v.ts::timestamp at time zone 'Asia/Manila'),",
        "       v.kind, 'biometric', 'intraday', 'synthetic'",
        "  from (values",
    ]
    out.append(",\n".join(
        f"    ({q(sim[p['name']])}, {q(p['ts'].strftime('%Y-%m-%d %H:%M:%S'))}, {q(p['kind'])})"
        for p in punches))
    out += [
        "  ) as v(full_name, ts, kind)",
        "  join faculty f on f.full_name = v.full_name",
        "  join faculty_pseudonym_map m on m.faculty_id = f.id;",
        "",
        "commit;",
        "",
    ]
    return "\n".join(out)


# ----------------------------------------------------------------------- xlsx
def _fit(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for c in ws[1]:
        c.font = Font(bold=True)
    ws.freeze_panes = "A2"


def emit_xlsx(path: pathlib.Path, punches, traits, calendar, names, blocks):
    sim = {n: sim_name(i) for i, n in enumerate(sorted(names), start=1)}
    pid = lambda real: pseudonym_for(sim[real])
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "README"
    lines = [
        ["ISU-GeoBot — synthetic attendance dataset"],
        [],
        ["THIS DATA IS INVENTED. It describes nobody."],
        [],
        ["Why it exists", "The thesis advisor ruled that real Daily Time Records cannot be"],
        ["", "requested, as a privacy concern, and recommended synthetic data."],
        ["", "Without attendance the Random Forest trains on schedule-derived"],
        ["", "features and schedule-derived labels, which reproduces the rule"],
        ["", "baseline by construction and cannot outperform it."],
        [],
        ["What it supports", "Measuring whether the pipeline RECOVERS the behavioural traits"],
        ["", "listed on the 'Faculty Traits' sheet, which were injected on purpose."],
        [],
        ["What it does NOT", "Any claim about real faculty availability, punctuality or"],
        ["support", "attendance at Isabela State University. No real person's"],
        ["", "behaviour informed any value in this file."],
        [],
        ["Reproducibility", f"Seed {SEED}. Regenerate with:"],
        ["", "python machine-learning/generate_synthetic_attendance.py"],
        [],
        ["Semester", f"{SEMESTER}  ({SEM_START} to {SEM_END})"],
        ["Lecturers", str(len(names))],
        ["Schedule blocks", f"{len(blocks)} (from the departmental workbook — these ARE real)"],
        ["Punch records", str(len(punches))],
        ["Granularity", "intraday (check_in / check_out pairs)"],
        ["data_origin", "synthetic — blocks reportable metrics by design"],
        [],
        ["Identifiers", "Rows are keyed on pseudonym_id, never a name. The pseudonym-to-"],
        ["", "name key is deliberately NOT in this workbook."],
        [],
        ["Calendar caveat", "The holidays and exam periods are plausible placeholders, not"],
        ["", "the ISU academic calendar. Replace them with the real one."],
    ]
    for row in lines:
        ws.append(row)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 78
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"].font = Font(bold=True)
    for r in ws.iter_rows(min_col=2, max_col=2):
        for c in r:
            c.alignment = Alignment(wrap_text=False, vertical="top")

    ws = wb.create_sheet("Punches")
    ws.append(["pseudonym_id", "date", "weekday", "event_time", "event_type",
               "source", "granularity", "data_origin"])
    wd = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    for p in punches:
        ws.append([pid(p["name"]), p["ts"].date().isoformat(),
                   wd[p["ts"].weekday()], p["ts"].strftime("%H:%M:%S"),
                   p["kind"], "biometric", "intraday", "synthetic"])
    _fit(ws, [22, 12, 9, 12, 12, 11, 12, 12])

    ws = wb.create_sheet("Faculty Traits")
    ws.append(["pseudonym_id", "punctuality_offset_min", "arrival_sd_min",
               "absence_rate", "early_departure_rate", "consultation_presence",
               "stay_late_min"])
    for n in sorted(names):
        t = traits[n]
        ws.append([pid(n), round(t.punctuality, 1), round(t.spread, 1),
                   round(t.absence, 4), round(t.early_departure, 3),
                   round(t.consultation, 3), round(t.stay_late, 1)])
    _fit(ws, [22, 22, 16, 14, 21, 22, 15])
    ws.append([])
    ws.append(["These are the injected ground truth — what the classifier must recover."])
    ws.append(["Negative punctuality_offset means the person tends to arrive early."])

    ws = wb.create_sheet("Calendar")
    ws.append(["date", "weekday", "event_type", "title", "disrupts_schedule", "data_origin"])
    for c in calendar:
        ws.append([c["date"].isoformat(), wd[c["date"].weekday()], c["type"],
                   c["title"], c["disrupts"], "synthetic"])
    _fit(ws, [12, 9, 16, 36, 18, 12])

    # Pseudonymised, like every other sheet.
    #
    # Names here would undo the pseudonymisation everywhere else. Synthetic
    # attendance tracks teaching hours closely, so a punch pattern read against
    # a NAMED timetable identifies whose record it is -- and the record is
    # invented. Pseudonymising the schedule keeps the sheet's usefulness (the
    # reviewer can still see the teaching load a pattern sits on) and removes
    # the join that makes it personal data.
    ws = wb.create_sheet("Schedule")
    ws.append(["pseudonym_id", "day", "start", "end", "course", "room", "campus", "data_origin"])
    days = {1: "Monday", 2: "Tuesday", 3: "Wednesday", 4: "Thursday", 5: "Friday"}
    for b in sorted(blocks, key=lambda x: (pid(x["faculty"]), x["dow"], x["start"])):
        ws.append([pid(b["faculty"]), days[b["dow"]], b["start"].strftime("%H:%M"),
                   b["end"].strftime("%H:%M"), b["course"], b["room"],
                   b["campus"], "real"])
    _fit(ws, [22, 11, 8, 8, 22, 12, 11, 12])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


# ------------------------------------------------------------------------ main
def main():
    ap = argparse.ArgumentParser(description="Generate synthetic intraday attendance")
    ap.add_argument("--workbook", type=pathlib.Path, default=si.WORKBOOK)
    ap.add_argument("--dry-run", action="store_true", help="report only; write nothing")
    ap.add_argument("--seed", type=int, default=SEED)
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.workbook, data_only=True)
    names, blocks = si.parse_faculty_sheet(wb["Faculty"])
    print(f"roster: {len(names)} lecturers, {len(blocks)} schedule blocks")

    rng = np.random.default_rng(args.seed)
    punches, traits, calendar = generate(blocks, names, rng)

    # ---- self-checks. Each of these has a specific failure it is guarding.
    pairs_ok = all(punches[i]["kind"] == "check_in" and punches[i + 1]["kind"] == "check_out"
                   and punches[i]["name"] == punches[i + 1]["name"]
                   and punches[i + 1]["ts"] > punches[i]["ts"]
                   for i in range(0, len(punches) - 1, 2))
    on_holiday = [p for p in punches if p["ts"].date() in HOLIDAYS]

    santiago_days = {(b["faculty"], b["dow"]) for b in blocks if b["campus"] == "santiago"}
    all_day_stgo = set()
    for f, d in santiago_days:
        local = [b for b in blocks
                 if b["faculty"] == f and b["dow"] == d and b["campus"] == "echague"]
        if not local:
            all_day_stgo.add((f, d))
    leaked = [p for p in punches
              if (p["name"], (p["ts"].weekday() + 1) % 7) in all_day_stgo]

    days = len({p["ts"].date() for p in punches})
    print(f"\n{len(punches):,} punches over {days} teaching days")
    print(f"  check_in/check_out pairing   {'ok' if pairs_ok else 'BROKEN'}")
    print(f"  punches on a holiday         {len(on_holiday)}")
    print(f"  punches on an all-Santiago day {len(leaked)}")
    print(f"  calendar rows                {len(calendar)}")

    per = {}
    for p in punches:
        per[p["name"]] = per.get(p["name"], 0) + 1
    lo = min(per.values()); hi = max(per.values())
    print(f"  punches per lecturer         {lo}..{hi}")

    # Anyone who teaches at Echague must appear. A lecturer silently reduced to
    # zero rows is invisible to the classifier and impossible to notice in a
    # 5,000-row table -- which is exactly what happened when Echague blocks
    # falling inside a travel window were dropped rather than started late.
    teaches_here = {b["faculty"] for b in blocks if b["campus"] == "echague"}
    missing = sorted(teaches_here - set(per))
    print(f"  lecturers teaching here      {len(teaches_here)}, "
          f"{len(missing)} with no punches")
    if missing:
        for m in missing:
            print(f"    MISSING: {m}")

    only_away = sorted({b["faculty"] for b in blocks} - teaches_here)
    if only_away:
        print(f"  Santiago-only (0 expected here): {', '.join(only_away)}")

    if not pairs_ok or on_holiday or leaked or missing:
        raise SystemExit("\nself-check failed -- refusing to write")

    if args.dry_run:
        print("\ndry run -- nothing written")
        return

    sql_path = ROOT / "database" / "sample-data" / "003_synthetic_attendance.sql"
    sql_path.parent.mkdir(parents=True, exist_ok=True)
    sql_path.write_text(emit_sql(punches, traits, calendar, names, blocks), encoding="utf-8")
    print(f"\nwrote {sql_path}")

    xlsx = ROOT / "machine-learning" / "training-data" / \
        "synthetic-attendance-1st-sem-2026-2027.xlsx"
    emit_xlsx(xlsx, punches, traits, calendar, names, blocks)
    print(f"wrote {xlsx}")


if __name__ == "__main__":
    main()
